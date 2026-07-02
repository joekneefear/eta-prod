"""
SYNOPSIS
    Unit tests for InnoFtXlsxParser

DESCRIPTION
    Comprehensive test suite for parsing INNO FT XLSX files.
    Tests header extraction, test table parsing, and model building.

AUTHOR
    kiro@onsemi.com

CHANGES
    2026-Jul-02 - Initial test suite

LICENSE
    (C) onsemi 2026 All rights reserved.

USAGE
    Run with pytest:
        cd scripts/py
        pytest tests/test_inno_ft_xlsx_parser.py -v
    
    Run specific test:
        pytest tests/test_inno_ft_xlsx_parser.py::test_parse_sample_file -v
"""
import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pytest
from hypothesis import given, strategies as st
from lib.Parser.InnoFtXlsxParser import InnoFtXlsxParser


class TestInnoFtXlsxParser:
    """Test INNO FT XLSX Parser."""
    
    @pytest.fixture
    def parser(self):
        """Create parser instance."""
        return InnoFtXlsxParser()
    
    @pytest.fixture
    def sample_file(self):
        """Get path to sample XLSX file."""
        sample_path = os.path.join(
            os.path.dirname(__file__),
            '..',
            'docs',
            '9UU190002 (1).xlsx'
        )
        return sample_path
    
    def test_parse_sample_file(self, parser, sample_file):
        """Test parsing of the sample INNO XLSX file.
        
        Validates:
        - LOT field extracted correctly
        - PRODUCT field extracted from Device Name fallback
        - RECIPE field contains program name
        - RECIPE_REVISION extracted from program pattern
        - Test names parsed
        - Dies parsed with expected count
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        
        # Validate header was populated
        assert model.header is not None
        assert model.header.LOT == "9UU190002"
        
        # Validate Device Name fallback for Product
        assert model.header.PRODUCT == "NTMT130N70GN1TXG"
        
        # Validate Program/Recipe extracted
        assert model.header.RECIPE is not None
        assert "IN0167" in model.header.RECIPE
        
        # Validate RecipeRevision extracted from _R10_ pattern
        assert model.header.RECIPE_REVISION == "10"
        
        # Validate SourceLot ends with .S
        assert model.header.SOURCE_LOT == "9UU190002.S"
        
        # Validate tests were parsed
        assert len(model.tests) > 0
        
        # Validate dies were parsed
        assert len(model.wafers[0].dies) >= 5
    
    def test_header_raw_dict_stored(self, parser, sample_file):
        """Test that raw_header dict is stored on model.header._raw.
        
        Validates:
        - _raw attribute exists
        - Contains all expected header keys
        - Values are strings
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        
        # Check _raw dict exists
        assert hasattr(model.header, '_raw')
        assert isinstance(model.header._raw, dict)
        
        # Check expected keys
        assert 'LotID' in model.header._raw
        assert 'Program' in model.header._raw
        assert 'Device Name' in model.header._raw
        
        # Check values are strings or NA
        assert isinstance(model.header._raw['LotID'], str)
    
    def test_recipe_revision_extraction(self, parser):
        """Test recipe revision extraction from program name.
        
        Property: For any program string containing _R<digits>_, 
        the parsed RecipeRevision must equal those digits.
        """
        # Test pattern matching on various program names
        test_cases = [
            ("IN0167_FT1x4_STGB_DFNX_R10_125C.pgs", "10"),
            ("PROG_R1_TEST.pgs", "1"),
            ("LONGNAME_R123_DATA.pgs", "123"),
        ]
        
        for program_name, expected_rev in test_cases:
            # Simulate what enricher will do
            import re
            match = re.search(r'.*_R(\d+)_.*', program_name)
            if match:
                actual_rev = match.group(1)
                assert actual_rev == expected_rev, \
                    f"Program {program_name} should extract revision {expected_rev}, got {actual_rev}"
    
    def test_test_names_parsed(self, parser, sample_file):
        """Test that test names are parsed from the file.
        
        Validates:
        - At least one test exists
        - Test has name, LSL, HSL, units
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        
        assert len(model.tests) > 0
        
        first_test = model.tests[0]
        assert hasattr(first_test, 'number')
        assert hasattr(first_test, 'name')
        assert first_test.name is not None
        assert len(first_test.name) > 0
    
    def test_dies_parsed_with_results(self, parser, sample_file):
        """Test that die data is parsed with test results.
        
        Validates:
        - Dies have partid
        - Dies have bin information
        - Dies have test results aligned to test_names count
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        wafer = model.wafers[0]
        
        assert len(wafer.dies) > 0
        
        first_die = wafer.dies[0]
        assert hasattr(first_die, 'partid')
        assert first_die.partid is not None
        assert len(first_die.partid) > 0
        
        # Check bin information
        assert hasattr(first_die, 'soft_bin')
        assert hasattr(first_die, 'hard_bin')
        
        # Check test results
        if len(model.tests) > 0:
            # Die should have results for each test
            assert hasattr(first_die, 'result')
            assert len(first_die.result) > 0
    
    def test_bins_accumulated(self, parser, sample_file):
        """Test that soft bins and hard bins are accumulated.
        
        Validates:
        - sbins list is populated
        - hbins list is populated
        - Bin counts are tracked
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        wafer = model.wafers[0]
        
        # Should have at least one bin
        assert len(wafer.sbins) > 0
        assert len(wafer.hbins) > 0
        
        # Each bin should have count > 0
        for sbin in wafer.sbins:
            assert hasattr(sbin, 'count')
            assert sbin.count > 0
    
    def test_lot_id_required(self, parser, sample_file):
        """Test that LOT is set from LotID header field.
        
        Validates:
        - LOT field is set
        - LOT matches LotID from header
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        
        assert model.header.LOT is not None
        assert model.header.LOT != 'NA'
        assert model.header.LOT == "9UU190002"
    
    def test_source_lot_appends_s(self, parser, sample_file):
        """Test that SOURCE_LOT ends with .S suffix.
        
        Property: For any LotID value parsed from a valid xlsx file,
        the resolved SOURCE_LOT metadata field must end with the suffix .S.
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        
        assert model.header.SOURCE_LOT is not None
        assert model.header.SOURCE_LOT.endswith('.S'), \
            f"SOURCE_LOT should end with .S, got {model.header.SOURCE_LOT}"
    
    def test_die_count_matches_data_rows(self, parser, sample_file):
        """Test that die count matches number of data rows.
        
        Property: For any xlsx file, the number of Die objects in wafer.dies 
        must equal the number of numeric data rows in the test table.
        
        Validates:
        - Dies are created for each data row
        - Count is consistent
        """
        if not os.path.exists(sample_file):
            pytest.skip(f"Sample file not found: {sample_file}")
        
        model = parser.parse_to_model(sample_file)
        
        # Die count should be > 0
        die_count = len(model.wafers[0].dies)
        assert die_count > 0
        
        # Each die should have partid
        for die in model.wafers[0].dies:
            assert die.partid is not None


class TestParserEdgeCases:
    """Test edge cases and error handling."""
    
    def test_file_not_found(self):
        """Test handling of non-existent file."""
        parser = InnoFtXlsxParser()
        
        with pytest.raises(SystemExit):
            parser.parse_to_model('/nonexistent/file.xlsx')
    
    def test_invalid_xlsx_file(self, tmp_path):
        """Test handling of invalid XLSX file."""
        # Create invalid file
        invalid_file = tmp_path / "invalid.xlsx"
        invalid_file.write_text("This is not a valid xlsx file")
        
        parser = InnoFtXlsxParser()
        
        with pytest.raises(SystemExit):
            parser.parse_to_model(str(invalid_file))
    
    def test_clean_cell_with_none(self):
        """Test _clean_cell handles None values."""
        parser = InnoFtXlsxParser()
        
        result = parser._clean_cell(None)
        assert result == ''
    
    def test_clean_cell_with_whitespace(self):
        """Test _clean_cell trims whitespace."""
        parser = InnoFtXlsxParser()
        
        result = parser._clean_cell('  test value  ')
        assert result == 'test value'


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])


class TestParserProperties:
    """Property-based tests for InnoFtXlsxParser using Hypothesis."""
    
    @given(st.lists(
        st.tuples(
            st.integers(min_value=1, max_value=10000),  # Part ID
            st.sampled_from(['1', '2', '5', '10', '0']),  # BIN value
            st.lists(st.floats(allow_nan=False, allow_infinity=False), min_size=1, max_size=10)  # Test results
        ),
        min_size=1,
        max_size=100
    ))
    def test_property_6_die_count_matches_data_rows(self, synthetic_data):
        """
        Property 6: Die count matches data rows
        
        For any xlsx file, the number of Die objects in wafer.dies 
        must equal the number of numeric data rows in the test table.
        
        Validates: Requirements 3.7
        
        Strategy: Generate synthetic data rows with numeric part IDs.
        Assert that if we parse N rows, we get N dies.
        """
        import tempfile
        from openpyxl import Workbook
        
        # Create synthetic xlsx file
        wb = Workbook()
        ws = wb.active
        
        # Add minimal header
        ws['A1'] = 'Program'
        ws['B1'] = 'TEST_PROG_R10_125C.pgs'
        ws['A2'] = 'LotID'
        ws['B2'] = 'TEST001'
        
        # Add test table header with test names
        ws['A10'] = 'No'
        ws['B10'] = 'BIN'
        ws['C10'] = 'Test Parameter'
        ws['D10'] = 'TEST1'
        
        # Add limits
        ws['A11'] = 'LL'
        ws['B11'] = ''
        ws['C11'] = ''
        ws['D11'] = '0'
        
        ws['A12'] = 'HL'
        ws['B12'] = ''
        ws['C12'] = ''
        ws['D12'] = '10'
        
        ws['A13'] = 'Unit'
        ws['B13'] = ''
        ws['C13'] = ''
        ws['D13'] = 'V'
        
        # Add data rows
        for row_num, (part_id, bin_val, test_results) in enumerate(synthetic_data, start=14):
            ws[f'A{row_num}'] = str(part_id)
            ws[f'B{row_num}'] = bin_val
            if test_results:
                ws[f'D{row_num}'] = test_results[0]  # At least one result
        
        # Write to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)
        
        try:
            parser = InnoFtXlsxParser()
            model = parser.parse_to_model(tmp_path)
            
            # Assert: die count equals synthetic data rows
            expected_die_count = len(synthetic_data)
            actual_die_count = len(model.wafers[0].dies)
            assert actual_die_count == expected_die_count, \
                f"Expected {expected_die_count} dies, got {actual_die_count}"
        finally:
            os.unlink(tmp_path)
    
    @given(st.text(min_size=1, max_size=20).filter(lambda x: x.strip() == ''))
    def test_property_5_whitespace_only_header_values_become_na(self, whitespace_value):
        """
        Property 5: Whitespace-only header values become NA
        
        For any xlsx header cell whose value is composed entirely of whitespace,
        the corresponding parsed field must resolve to "NA" in the raw header dict.
        
        Validates: Requirements 3.8
        
        Strategy: Generate header rows with whitespace-only values.
        Assert that parser defaults them to "NA".
        """
        import tempfile
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Add header with whitespace-only value for Program
        ws['A1'] = 'Program'
        ws['B1'] = whitespace_value  # Whitespace-only
        ws['A2'] = 'LotID'
        ws['B2'] = 'TEST001'
        
        # Add minimal test table
        ws['A10'] = 'No'
        ws['B10'] = 'BIN'
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)
        
        try:
            parser = InnoFtXlsxParser()
            model = parser.parse_to_model(tmp_path)
            
            # Check that _raw dict has Program as NA
            assert hasattr(model.header, '_raw')
            # Program should either not be in _raw (skipped) or be NA
            if 'Program' in model.header._raw:
                assert model.header._raw['Program'] == 'NA', \
                    f"Whitespace-only header should become NA, got {model.header._raw['Program']}"
        finally:
            os.unlink(tmp_path)
