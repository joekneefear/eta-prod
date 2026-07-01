"""
SYNOPSIS

DESCRIPTION
    Parser for SHEDCL DTS1000/DTS2000 (JUNO) Excel format test data
    Converted from Perl module PDF::Parser::Dts_2k_xls
  
AUTHOR
    junifferallan.garcia@onsemi.com

CHANGES
    2026-Jan-29 - Initial Python implementation with custom extractor support
    2026-Jan-29 - Optimized for performance and best practices

LICENSE
    (C) onsemi 2026 All rights reserved.
"""



import os
import re
from typing import Optional, List, Dict, Any, Tuple
from collections import defaultdict
from openpyxl import load_workbook

from lib.Data.Model import Model
from lib.Data.Wafer import Wafer
from lib.Data.Test import Test
from lib.Data.Die import Die
from lib.Data.Bin import Bin
from lib.Data.Metadata import Metadata
from lib.Log import Log
from lib.Util import Util
from lib.Config.Dts1k2kParserConfig import Dts1k2kParserConfig


class Dts1k2kXlsParser:
    """
    Parser for SHEDCL DTS1000/DTS2000 (JUNO) Excel format test data.
    
    Supports custom field extraction through ParserConfig for flexible
    parsing of lot IDs, test programs, timestamps, and other fields.
    
    Optimized for performance with compiled regex patterns and efficient
    data structures.
    """
    
    # Compile regex patterns once at class level for performance
    _PATTERNS = {
        'date': re.compile(r'Date', re.IGNORECASE),
        'version': re.compile(r'Version', re.IGNORECASE),
        'station': re.compile(r'Station', re.IGNORECASE),
        'systemname': re.compile(r'SystemName', re.IGNORECASE),
        'devicename': re.compile(r'Device(?:Name)?$', re.IGNORECASE),  # Matches 'Device' or 'DeviceName'
        'lotname': re.compile(r'Lot(?:Name)?$', re.IGNORECASE),  # Matches 'Lot' or 'LotName'
        'operatorname': re.compile(r'Operator(?:Name)?$', re.IGNORECASE),  # Matches 'Operator' or 'OperatorName'
        'testfilename': re.compile(r'TestFileName', re.IGNORECASE),
        'item_name': re.compile(r'Item\s+Name|Item_Name', re.IGNORECASE),
        'bias': re.compile(r'Bias[123]', re.IGNORECASE),
        'min_limit': re.compile(r'Min_Limit|Min\sLimit', re.IGNORECASE),
        'max_limit': re.compile(r'Max_Limit|Max\sLimit', re.IGNORECASE),
        'serial': re.compile(r'Serial', re.IGNORECASE),
        'bin': re.compile(r'Bin', re.IGNORECASE),
        'sortbin': re.compile(r'^SortBin$'),
        'data_row_numeric': re.compile(r'^\d+$'),
        'data_row_pf': re.compile(r'^[PF]\d+$'),
        'non_digit': re.compile(r'\D'),
        'non_digit_end': re.compile(r'\D+$'),
        'non_digit_edges': re.compile(r'^\D+|\D+$'),
        'whitespace': re.compile(r'\s+'),
        'extension': re.compile(r'\..+$'),
        'tp_revision': re.compile(r'V(\d+)$', re.IGNORECASE),
        'digit_dash_dot': re.compile(r'[\d\-\.]'),
        # Extracts leading number from item name: "2 ISGS" -> num=2, name=ISGS
        'test_num_prefix': re.compile(r'^(\d+)[\s:_]+', re.IGNORECASE),
        # Parses a bias cell value like "IB=1.00mA" -> groups: label, numeric_value, units
        'bias_cell': re.compile(r'^([A-Za-z_]+\d*)\s*=\s*([\d\.eE+\-]+)\s*([A-Za-z%µu/]*)\s*$'),
    }
    
    def __init__(self, config: Optional[Dts1k2kParserConfig] = None, pplogger=None):
        """
        Initialize parser with optional custom configuration and pplogger.
        
        Args:
            config: Dts1k2kParserConfig with custom extractors, or None for default parsing
            pplogger: PPLogger instance for logging to refdb.pp_log (optional)
        """
        self.config = config or Dts1k2kParserConfig()
        self.pplogger = pplogger
        self.logger = Log()
        
        # Cache extractor lookups for performance
        self._extractor_cache: Dict[str, Optional[Any]] = {}
    
    def parse_to_model(self, infile: str, original_filename: Optional[str] = None) -> Model:
        """
        Parse Excel file into Model object.
        
        Args:
            infile: Path to Excel file (.xls or .xlsx) to read content from
            original_filename: Original filename (before decompression) for metadata extraction
            
        Returns:
            Model object with parsed data
            
        Raises:
            FileNotFoundError: If input file doesn't exist
            Exception: If Excel file is invalid or parsing fails
        """
        if not os.path.exists(infile):
            error_msg = f"File not found: {infile}"
            self.logger.ERROR(error_msg)
            Util.dp_exit(1, pplogger=self.pplogger, error=error_msg)
        
        self.logger.INFO(f"Parsing DTS1000/DTS2000 file: {infile}")
        
        # Initialize model
        header = Metadata()
        model = Model({
            'header': header,
            'misc': {},
            'dataSource': 'JUNO_DTS1000/DTS2000'
        })
        
        # Initialize wafer (wafer number 0 for discrete device testing)
        wafer = Wafer({'number': 0})
        model.add('wafers', wafer)
        
        # Parse Excel file - use original filename for metadata extraction if provided
        metadata_source = original_filename if original_filename else infile
        context = {'input_file': metadata_source}
        
        self.logger.INFO(f"Using filename for metadata extraction: {metadata_source}")
        
        # Proactively apply extractors if they are configured to use filename/external sources
        # This ensures metadata is populated even if markers (LotName, Date, etc.) are missing in the file
        self._apply_proactive_extractors(header, context)
        
        self._parse_excel_file(infile, model, wafer, header, context)
        
        # Log parsed metadata
        self.logger.INFO(f"LOT={header.LOT}--DEVICE={header.PRODUCT}--PROGRAM={header.PROGRAM}--RECIPE_REVISION={header.RECIPE_REVISION}--TIME={header.START_TIME}")
        
        # Validate Lot ID
        if not header.LOT or header.LOT == 'NA':
            error_msg = f"BAD FILE FORMAT - NO valid Lotid={header.LOT}!"
            self.logger.ERROR(error_msg)
            Util.dp_exit(1, pplogger=self.pplogger, error=error_msg)
            
        # Build limits from tests
        model.build_limit()
        model.limit.testItems = ['number', 'name', 'units']
        
        self.logger.INFO(
            f"Parsing complete: {len(model.tests)} tests, {len(wafer.dies)} dies, "
            f"{len(wafer.sbins)} sbins, {len(wafer.hbins)} hbins"
        )
        
        return model
    def _apply_proactive_extractors(self, header: Metadata, context: Dict[str, Any]) -> None:
        """
        Apply extractors that can work from external info (filename, mtime).
        Only applies extractors that don't require file content.
        
        Args:
            header: Metadata object to populate
            context: Parser context
        """
        input_file = context.get('input_file', '')
        
        # Check if lot extractor uses external source (filename)
        # Skip if it uses 'content' source since we need to read the file first
        lot_config = self.config.config_data.get('lot_id', {})
        lot_source = lot_config.get('source', 'content')
        
        if lot_source == 'filename':
            # Apply Lot/Device extractor from filename
            extractor = self._get_extractor('lot_parser')
            if extractor:
                result = extractor({'LotName': ''}, context)
                if result:
                    for key, val in result.items():
                        if val and val != 'NA':
                            old_val = getattr(header, key, None)
                            if old_val != val:
                                setattr(header, key, val)
                                self.logger.INFO(f"Proactively extracted {key}={val} from filename")
        else:
            self.logger.INFO(f"Skipping proactive lot extraction (source={lot_source}, requires file content)")
                    
        # Apply Time extractor (for START/END_TIME from mtime)
        extractor = self._get_extractor('time_parser')
        if extractor:
            result = extractor({}, context)
            if result:
                for key, val in result.items():
                    if val and val != 'NA':
                        setattr(header, key, val)
                        self.logger.INFO(f"Proactively extracted {key}={val} from file system")
                    
        # Check if program extractor uses external source
        program_config = self.config.config_data.get('program', {})
        program_source = program_config.get('source', 'content')
        
        # Only apply program extractor if it's configured to use filename
        if program_source == 'filename':
            extractor = self._get_extractor('program_parser')
            if extractor:
                result = extractor({'TestFileName': input_file}, context)
                if result:
                    for key, val in result.items():
                        if val and val != 'NA':
                            setattr(header, key, val)
                            self.logger.INFO(f"Proactively extracted {key}={val} from filename")
        else:
            self.logger.INFO(f"Skipping proactive program extraction (will use file content)")
    
    def _get_extractor(self, extractor_name: str) -> Optional[Any]:
        """
        Get extractor with caching for performance.
        
        Args:
            extractor_name: Name of the extractor
            
        Returns:
            Extractor function or None
        """
        if extractor_name not in self._extractor_cache:
            self._extractor_cache[extractor_name] = (
                self.config.get_extractor(extractor_name) 
                if self.config.has_extractor(extractor_name) 
                else None
            )
        return self._extractor_cache[extractor_name]
    
    def _load_file_optimized(self, infile: str) -> Tuple[Any, int]:
        """
        Load file using the most efficient library for the format.
        
        Automatically detects file type and uses:
        - xlrd for .xls files (3-5x faster than openpyxl)
        - pandas for .csv files (fastest)
        - openpyxl for .xlsx files (optimized)
        
        Args:
            infile: Path to file
            
        Returns:
            Tuple of (row_iterator, row_count)
        """
        file_ext = os.path.splitext(infile)[1].lower()
        
        if file_ext == '.csv':
            return self._load_csv(infile)
        elif file_ext == '.xls':
            return self._load_xls(infile)
        elif file_ext in ['.xlsx', '.xlsm']:
            return self._load_xlsx(infile)
        else:
            # Try openpyxl as fallback
            self.logger.WARN(f"Unknown file extension {file_ext}, trying openpyxl")
            return self._load_xlsx(infile)
    
    def _load_csv(self, infile: str) -> Tuple[Any, int]:
        """Load CSV file using pandas (fastest) or csv module as fallback."""
        try:
            import pandas as pd
            try:
                # Use header=None to avoid column count detection issues with metadata headers
                df = pd.read_csv(infile, header=None)
            except (UnicodeDecodeError, pd.errors.ParserError):
                self.logger.WARN(f"Pandas load failed, retrying with latin1 and header=None: {infile}")
                df = pd.read_csv(infile, header=None, encoding='latin1')
            
            self.logger.INFO(f"Loaded CSV with pandas: {len(df)} rows")
            # Convert to list of lists for consistent interface
            return (list(row) for row in df.values), len(df)
        except Exception as e:
            # Catch everything else (missing pandas, persistent ParserError) and fallback to csv module
            self.logger.WARN(f"Pandas failed completely ({e}), falling back to robust csv module")
            import csv
            # Fallback to csv module with robust encoding error handling
            with open(infile, 'r', encoding='latin1', errors='replace') as f:
                reader = csv.reader(f)
                rows = list(reader)
            self.logger.INFO(f"Loaded CSV with csv module: {len(rows)} rows")
            return iter(rows), len(rows)
    
    def _load_xls(self, infile: str) -> Tuple[Any, int]:
        """Load .xls file using xlrd (3-5x faster than openpyxl) or fallback to openpyxl."""
        try:
            import xlrd
            workbook = xlrd.open_workbook(infile, on_demand=True)
            worksheet = workbook.sheet_by_index(0)
            
            def row_generator():
                for row_idx in range(worksheet.nrows):
                    yield worksheet.row_values(row_idx)
            
            self.logger.INFO(f"Loaded XLS with xlrd: {worksheet.nrows} rows (optimized)")
            return row_generator(), worksheet.nrows
        except ImportError:
            self.logger.WARN("xlrd not installed, falling back to openpyxl (slower for .xls files)")
            return self._load_xlsx(infile)
        except Exception as e:
            self.logger.WARN(f"xlrd failed ({e}), falling back to openpyxl")
            return self._load_xlsx(infile)
    
    def _load_xlsx(self, infile: str) -> Tuple[Any, int]:
        """Load .xlsx file using openpyxl with optimizations."""
        try:
            workbook = load_workbook(infile, data_only=True, read_only=True)
            worksheet = workbook.worksheets[0]
            
            # Use max_row to avoid iterating through empty rows
            max_row = worksheet.max_row or 1000000
            
            self.logger.INFO(f"Loaded XLSX with openpyxl: ~{max_row} rows")
            
            # Return iterator with max_row limit
            return worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True), max_row
        except Exception as e:
            error_msg = f"Failed to load Excel file: {e}"
            self.logger.ERROR(error_msg)
            Util.dp_exit(1, pplogger=self.pplogger, error=error_msg)
    
    def _parse_excel_file(self, infile: str, model: Model, wafer: Wafer, 
                          header: Metadata, context: Dict[str, Any]) -> None:
        """
        Parse file and populate model.
        
        Uses optimized file loading based on format:
        - .xls: xlrd (3-5x faster)
        - .csv: pandas or csv module
        - .xlsx: openpyxl (read-only mode)
        
        Args:
            infile: Path to file
            model: Model object to populate
            wafer: Wafer object to populate
            header: Metadata object to populate
            context: Parser context for custom extractors
        """
        # Load file with optimal library for format
        row_iterator, estimated_rows = self._load_file_optimized(infile)
        
        # Parsing state - use defaultdict for bin_counts
        test_names: List[str] = []
        hi_limits: List[str] = []
        lo_limits: List[str] = []
        hi_units: List[str] = []
        lo_units: List[str] = []
        # bias_rows: list of per-bias-row lists, each aligned to test columns
        # e.g. bias_rows[0] = ["IB=1.00mA", "IC=5.0mA", ...] for Bias1
        bias_rows: List[List[str]] = []
        data_flag = False
        sortbin_flag = False
        
        # Store station for equipment ID construction
        station = ""
        systemname = ""
        
        # Parse rows with optimized iteration
        for row in row_iterator:
            # Skip empty rows early (before list comprehension)
            if not row or not row[0]:
                continue
            
            # Convert row to list of cleaned strings
            row_data = [self._clean_string(cell) for cell in row]
            
            field_name = row_data[0]
            field_value = row_data[2] if len(row_data) > 2 else ''
            
            # Use compiled patterns for faster matching
            # Parse metadata rows
            if self._PATTERNS['date'].match(field_name):
                self._parse_date_field(row_data, header, context)
                
            elif self._PATTERNS['version'].match(field_name):
                self._parse_version_field(row_data, header, context)
                
            elif self._PATTERNS['station'].match(field_name):
                station = Util.trim(field_value)
                
            elif self._PATTERNS['systemname'].match(field_name):
                systemname = Util.trim(field_value)
                header.EQUIP1_ID = f"{systemname} {station}"
                
            elif self._PATTERNS['devicename'].match(field_name):
                self._parse_device_field(row_data, header, context)
                
            elif self._PATTERNS['lotname'].match(field_name):
                self._parse_lot_field(row_data, header, context)
                
            elif self._PATTERNS['operatorname'].match(field_name):
                header.OPERATOR = Util.trim(field_value)
                
            elif self._PATTERNS['testfilename'].match(field_name):
                self._parse_program_field(row_data, header, context)
            
            # Parse test definition rows
            elif (len(row_data) > 1 and self._PATTERNS['item_name'].match(row_data[1])) or \
                 (self._PATTERNS['item_name'].match(field_name) and (len(row_data) < 2 or not row_data[1])):
                # Test names row
                raw_test_names = row_data[2:]
                self.logger.INFO(f"Raw test names count: {len(raw_test_names)}")
                
                # Filter out empty test names (caused by trailing commas in CSV)
                test_names = [name for name in raw_test_names if name and name.strip()]
                self.logger.INFO(f"Filtered test names count: {len(test_names)} (removed {len(raw_test_names) - len(test_names)} empty columns)")
                
                # Check if last column is SortBin
                if test_names and self._PATTERNS['sortbin'].match(test_names[-1]):
                    test_names.pop()  # Remove SortBin from test names
                    sortbin_flag = True
                    self.logger.INFO(f"Removed SortBin column, final test count: {len(test_names)}")
                
                self.logger.INFO(f"Final test names: {test_names[:5]}..." if len(test_names) > 5 else f"Final test names: {test_names}")
                    
            elif self._PATTERNS['bias'].match(field_name):
                self._parse_bias_row(row_data, bias_rows, sortbin_flag)
                
            elif self._PATTERNS['min_limit'].match(field_name):
                self._parse_limit_row(row_data, lo_limits, lo_units, sortbin_flag)
                
            elif self._PATTERNS['max_limit'].match(field_name):
                self._parse_limit_row(row_data, hi_limits, hi_units, sortbin_flag)
                
            elif self._PATTERNS['serial'].match(field_name) and \
                 len(row_data) > 1 and self._PATTERNS['bin'].match(row_data[1]):
                # Data header row
                data_flag = True
                
            elif data_flag and self._is_data_row(row_data):
                # Parse die data
                self._parse_die_data(row_data, wafer, test_names)
        
        # Create test objects
        self._create_tests(model, wafer, test_names, lo_limits, hi_limits, 
                          lo_units, hi_units, bias_rows)
    
    def _clean_string(self, value: Any) -> str:
        """
        Clean cell value: strip whitespace, remove commas, replace spaces with underscores.
        
        Args:
            value: Cell value from Excel
            
        Returns:
            Cleaned string
        """
        # Handle None and NaN (from pandas)
        if value is None or str(value).lower() == 'nan':
            return ''
        
        # Chain operations for efficiency
        str_val = str(value).strip().replace(',', '')
        str_val = self._PATTERNS['whitespace'].sub('_', str_val)
        
        return str_val
    
    def _parse_date_field(self, row_data: List[str], header: Metadata, context: Dict[str, Any]) -> None:
        """Parse Date field to extract START_TIME and END_TIME."""
        # Check if custom time parser is registered (cached lookup)
        extractor = self._get_extractor('time_parser')
        if extractor:
            result = extractor({}, context)
            if result:
                for key, val in result.items():
                    if val and val != 'NA':
                        setattr(header, key, val)
                return
        
        # Default parsing: Date row format is "Date", "", "Start_<date>_<time>", "", "", "End_<date>_<time>"
        if len(row_data) > 2:
            parts = row_data[2].split('_')
            if len(parts) >= 3:
                val = f"{parts[1]} {parts[2]}"
                if val and val != 'NA':
                    header.START_TIME = val
        
        if len(row_data) > 5:
            parts = row_data[5].split('_')
            if len(parts) >= 3:
                val = f"{parts[1]} {parts[2]}"
                if val and val != 'NA':
                    header.END_TIME = val
    
    def _parse_version_field(self, row_data: List[str], header: Metadata, context: Dict[str, Any]) -> None:
        """Parse Version field."""
        if len(row_data) > 2:
            header.RECIPE_REVISION = Util.trim(row_data[2])
    
    def _parse_device_field(self, row_data: List[str], header: Metadata, context: Dict[str, Any]) -> None:
        """Parse DeviceName field."""
        if len(row_data) > 2:
            device_value = Util.trim(row_data[2])
            if not device_value or device_value == 'NA':
                return
            
            # Check for custom device parser (cached lookup)
            extractor = self._get_extractor('device_parser')
            if extractor:
                result = extractor({'DeviceName': device_value}, context)
                for key, val in result.items():
                    if val and val != 'NA':
                        setattr(header, key, val)
            else:
                header.PRODUCT = device_value
    
    def _parse_lot_field(self, row_data: List[str], header: Metadata, context: Dict[str, Any]) -> None:
        """Parse Lot/LotName field with optional custom extractor."""
        if len(row_data) > 2:
            lot_value = Util.trim(row_data[2])
            self.logger.INFO(f"Found Lot field with value: '{lot_value}'")
            
            if not lot_value or lot_value == 'NA':
                self.logger.WARN(f"Lot field is empty or NA, skipping extraction")
                return
            
            # Check for custom lot parser (cached lookup)
            extractor = self._get_extractor('lot_parser')
            if extractor:
                self.logger.INFO(f"Applying custom lot extractor to value: '{lot_value}'")
                result = extractor({'LotName': lot_value}, context)
                for key, val in result.items():
                    if val and val != 'NA':
                        setattr(header, key, val)
                        self.logger.INFO(f"Extracted from Lot field: {key}={val}")
            else:
                header.LOT = lot_value
                self.logger.INFO(f"No custom extractor, set LOT={lot_value}")
    
    def _parse_program_field(self, row_data: List[str], header: Metadata, context: Dict[str, Any]) -> None:
        """Parse TestFileName field with optional custom extractor."""
        if len(row_data) > 2:
            test_filename = row_data[2]
            if not test_filename or test_filename == 'NA':
                return
            
            # Check for custom program parser (cached lookup)
            extractor = self._get_extractor('program_parser')
            if extractor:
                result = extractor({'TestFileName': test_filename}, context)
                for key, val in result.items():
                    if val and val != 'NA':
                        setattr(header, key, val)
            else:
                # Mirror Perl's behavior (Juno_Data_xls.pm):
                # 1. Robustly get filename (handle both / and \ regardless of OS)
                basename = re.split(r'[\\/]', test_filename)[-1]
                basename = self._PATTERNS['extension'].sub('', basename)
                
                # Extract revision if V<digit> pattern exists at the end
                rev_match = self._PATTERNS['tp_revision'].search(basename)
                if rev_match:
                    revision = rev_match.group(1)
                    basename = self._PATTERNS['tp_revision'].sub('', basename)
                    header.RECIPE_REVISION = revision
                    self.logger.INFO(f"Extracted Revision from filename: {revision}")
                
                if basename and basename != 'NA':
                    header.PROGRAM = basename
    
    def _parse_bias_row(self, row_data: List[str], bias_rows: List[List[str]], sortbin_flag: bool) -> None:
        """
        Collect one Bias row (Bias1, Bias2, or Bias3) as a per-column list.
        Each cell is kept as-is (e.g. "IB=1.00mA") so _create_tests can build
        the composite test name per-column using only the defined (non-empty) biases.
        """
        items = row_data[2:]
        if sortbin_flag and items:
            items = items[:-1]

        row_cells: List[str] = []
        for item in items:
            if not item or not item.strip() or item.lower() == 'undef':
                row_cells.append('')
            else:
                row_cells.append(item.strip())

        bias_rows.append(row_cells)
    

    def _parse_limit_row(self, row_data: List[str], limits: List[str], units: List[str], sortbin_flag: bool) -> None:
        """Parse Min_Limit or Max_Limit row. Extracts numeric value and unit separately."""
        items = row_data[2:]
        if sortbin_flag and items:
            items = items[:-1]

        for item in items:
            if not item or not item.strip():
                limits.append('')
                units.append('')
                continue

            # Extract numeric limit value (strip trailing non-digit chars)
            limit_val = self._PATTERNS['non_digit_end'].sub('', item)
            # Extract unit (strip all digit/dash/dot chars)
            unit_val = self._PATTERNS['digit_dash_dot'].sub('', item)

            limits.append(limit_val)
            units.append(unit_val)

    def _is_data_row(self, row_data: List[str]) -> bool:
        """Check if row is a data row (Serial# column starts with digits or P/F+digits)."""
        if not row_data or not row_data[0]:
            return False
        first_cell = row_data[0]
        return bool(
            self._PATTERNS['data_row_numeric'].match(first_cell) or
            self._PATTERNS['data_row_pf'].match(first_cell)
        )

    def _parse_die_data(self, row_data: List[str], wafer: Wafer, test_names: List[str]) -> None:
        """
        Parse one data row into a Die object.

        Column mapping per spec:
          col[0] = Serial# -> partid (numeric only)
          col[1] = Bin#    -> soft_bin and hard_bin (numeric only)
          col[2+]          = test results aligned to test_names
        """
        if len(row_data) < 2:
            return

        raw_partid = Util.trim(row_data[0])
        raw_bin    = Util.trim(row_data[1])

        # P/F prefix on serial indicates pass/fail directly
        pass_fail = raw_partid[0] if raw_partid and raw_partid[0] in ('P', 'F') else None

        # Strip non-digits to get numeric partid and bin
        partid   = self._PATTERNS['non_digit'].sub('', raw_partid) or raw_partid
        soft_bin = self._PATTERNS['non_digit'].sub('', raw_bin) or '0'

        if pass_fail is None:
            pass_fail = 'P' if soft_bin == '1' else 'F'

        die = Die({
            'partid':        partid,
            'soft_bin':      soft_bin,
            'hard_bin':      soft_bin,
            'bindesc':       f"SWBin_{soft_bin.zfill(3)}",
            'site':          1,
            'touchdown_num': -1,
            'ecid':          partid,
        })

        # Parse test results aligned to test_names
        for result_val in row_data[2:len(test_names) + 2]:
            cleaned = result_val.replace('Over', '').replace('undef', '')
            cleaned = self._PATTERNS['non_digit_edges'].sub('', cleaned)
            cleaned = Util.trim(cleaned)
            die.add('result', Util.rep_na(cleaned))

        wafer.add('dies', die)

        # Track sbins
        bin_name_s = f"SWBin_{soft_bin.zfill(3)}"
        bin_obj_s  = wafer.find("sbins", {"number": soft_bin})
        if bin_obj_s is None:
            wafer.add("sbins", Bin({
                'number':  soft_bin,
                'name':    bin_name_s,
                'bindesc': bin_name_s,
                'PF':      pass_fail,
                'count':   1,
            }))
        else:
            bin_obj_s.count += 1

        # Track hbins
        bin_name_h = f"HWBin_{soft_bin.zfill(3)}"
        bin_obj_h  = wafer.find("hbins", {"number": soft_bin})
        if bin_obj_h is None:
            wafer.add("hbins", Bin({
                'number':  soft_bin,
                'name':    bin_name_h,
                'bindesc': bin_name_h,
                'PF':      pass_fail,
                'count':   1,
            }))
        else:
            bin_obj_h.count += 1

    def _build_test_name_with_bias(self, base_name: str, col_idx: int, bias_rows: List[List[str]]) -> str:
        """
        Build the full test name for a column using the formula:
            <base_name>_<BiasLabel>=<BiasValue><BiasUnits>[_<Bias2Label>=...]
        Only bias rows that have a non-empty value for this column are included.

        Bias cells arrive as already-cleaned strings like "IB=1.00mA" or
        "IB_=_1.00_mA" (after _clean_string ran whitespace->_ substitution).
        We normalise them back to "IB=1.00mA" form for the name.
        """
        parts = [base_name]
        for row_cells in bias_rows:
            if col_idx >= len(row_cells):
                continue
            cell = row_cells[col_idx]
            if not cell:
                continue
            # Normalise: collapse any underscores around '=' that _clean_string may have introduced
            # e.g. "IB_=_1.00_mA" -> "IB=1.00mA"
            cell = re.sub(r'\s*_*=_*\s*', '=', cell)   # fix "_=_" around equals
            cell = re.sub(r'_+', '_', cell)             # collapse multiple underscores
            cell = cell.strip('_')
            if cell:
                parts.append(cell)
        return '_'.join(parts)

    def _create_tests(self, model: Model, wafer: Wafer, test_names: List[str],
                      lo_limits: List[str], hi_limits: List[str],
                      lo_units: List[str], hi_units: List[str],
                      bias_rows: List[List[str]]) -> None:
        """
        Create Test objects from parsed header rows.

        Test number  : digit prefix extracted from Item Name column value
                       e.g. "2 BVCEO" -> number=2
        Test name    : base name from Item Name + bias conditions appended per formula:
                       <ItemName>_<B1Label>=<B1Val><B1Units>[_<B2Label>=<B2Val><B2Units>...]
                       Only biases that are defined (non-empty) for the column are included.
        Units        : from Min/Max Limit rows (underscores removed).
        LSL          : Min Limit numeric value.
        HSL          : Max Limit numeric value.
        Bias fields  : used only for name derivation; NOT stored as separate Test attributes.
        """
        self.logger.INFO(
            f"Creating tests: {len(test_names)} names, {len(lo_limits)} lo_limits, "
            f"{len(hi_limits)} hi_limits, {len(bias_rows)} bias rows"
        )

        for i, raw_name in enumerate(test_names):
            lsl     = lo_limits[i] if i < len(lo_limits) else ''
            hsl     = hi_limits[i] if i < len(hi_limits) else ''
            lo_unit = lo_units[i]  if i < len(lo_units)  else ''
            hi_unit = hi_units[i]  if i < len(hi_units)  else ''

            # --- Extract test number from leading digits in Item Name ---
            num_match = self._PATTERNS['test_num_prefix'].match(raw_name)
            test_number = int(num_match.group(1)) if num_match else (i + 1)

            # Base name = item name with number prefix stripped + underscore artifacts cleaned
            base_name = self._PATTERNS['test_num_prefix'].sub('', raw_name).strip('_')

            # --- Build full test name with bias conditions ---
            full_name = self._build_test_name_with_bias(base_name, i, bias_rows)

            # --- Unit consistency check ---
            if lo_unit and hi_unit and lo_unit != hi_unit:
                Util.dp_exit(
                    1, pplogger=self.pplogger,
                    error=f"Inconsistent unit for test '{raw_name}': "
                          f"Min={lo_unit} vs Max={hi_unit}"
                )

            raw_unit = lo_unit or hi_unit or ''

            test = Test()
            test.number = test_number
            test.name   = full_name
            test.units  = raw_unit.replace('_', '').strip()
            test.LSL    = lsl
            test.HSL    = hsl

            # Bias fields are embedded in the name only; no separate conditions stored

            model.add('tests', test)
            wafer.add('tests', test)

    def _create_bins(self, wafer: Wafer, bin_counts: Dict[int, int]) -> None:
        """Create Bin objects from bin counts (legacy helper, not used in main flow)."""
        for bin_num, count in bin_counts.items():
            sbin = wafer.find('bins', {'number': bin_num})
            if not sbin:
                sbin = Bin()
                wafer.add('bins', sbin)
            sbin.number = bin_num
            sbin.name   = f"BIN_{bin_num}"
            sbin.PF     = "P" if bin_num == 1 else "F"
            sbin.count  = count
