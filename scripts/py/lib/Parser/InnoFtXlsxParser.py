"""
SYNOPSIS
    Parser for INNO Final Test (FT) Excel format test data

DESCRIPTION
    Parses .xlsx files produced by INNO test equipment.
    Extracts header metadata and test results into a Model object.
    Header block contains metadata labels (Program, Product, etc.)
    Test table contains test names, limits, units, and die data.

AUTHOR
    junifferallan.garcia@onsemi.com

CHANGES
    2026-Jul-02 - Initial Python implementation
    2026-Jul-03 - Fixed test result alignment and parsing logic to handle test headers
                  appearing before the "No" data marker

LICENSE
    (C) onsemi 2026 All rights reserved.
"""

import os
import re
from typing import Optional, List, Dict, Any, Tuple
from openpyxl import load_workbook

from lib.Data.Model import Model
from lib.Data.Wafer import Wafer
from lib.Data.Test import Test
from lib.Data.Die import Die
from lib.Data.Bin import Bin
from lib.Data.Metadata import Metadata
from lib.Log import Log
from lib.Util import Util


class InnoFtXlsxParser:
    """
    Parser for INNO Final Test (FT) Excel format.

    Parses xlsx files with:
    - Header block: key-value pairs (label in col A, value in col B or C)
    - Test headers: test names, limits (LL/HL), units (BEFORE "No" marker)
    - Die data rows (AFTER "No" marker)

    Outputs a Model with:
    - model.header: Metadata with parsed fields (raw_header stored as _raw attr)
    - model.tests: List of Test objects with limits and units
    - model.wafers[0].dies: List of Die objects with test results
    - model.wafers[0].sbins/hbins: Bin objects with counts
    """

    # Compiled regex patterns for performance
    _PATTERNS = {
        'numeric_row': re.compile(r'^\d+$'),
        'test_param': re.compile(r'Test\s*Parameter', re.IGNORECASE),
        'll_limit': re.compile(r'^LL$', re.IGNORECASE),
        'hl_limit': re.compile(r'^HL$', re.IGNORECASE),
        'unit_row': re.compile(r'^Unit$', re.IGNORECASE),
        'test_table_marker': re.compile(r'^No$', re.IGNORECASE),
        'test_num_row': re.compile(r'Test\s*#', re.IGNORECASE),
        'whitespace': re.compile(r'^\s*$'),
    }

    # Known header labels to capture
    _HEADER_LABELS = {
        'Program', 'Product', 'WaferModle', 'LotID', 'TesterId', 'Handler',
        'Device Name', 'Test temp', 'TestDate', 'Sub LotID', 'Operator ID'
    }

    def __init__(self, pplogger=None):
        """
        Initialize parser.

        Args:
            pplogger: Optional PPLogger instance for logging (not used in this phase)
        """
        self.pplogger = pplogger
        self.logger = Log()

    def parse_to_model(self, infile: str) -> Model:
        """
        Parse Excel file into Model object.

        Args:
            infile: Path to .xlsx file to parse

        Returns:
            Model object with parsed data

        Raises:
            FileNotFoundError: If input file doesn't exist
            Exception: If Excel file is invalid or parsing fails
        """
        if not os.path.exists(infile):
            error_msg = f"File not found: {infile}"
            self.logger.ERROR(error_msg)
            Util.dp_exit(1, pplogger=self.pplogger, error=error_msg)

        self.logger.INFO(f"Parsing INNO FT XLSX file: {infile}")

        # Initialize model
        header = Metadata()
        model = Model({
            'header': header,
            'misc': {},
            'dataSource': 'INNO_FT_XLSX'
        })

        # Initialize wafer (wafer number 0 for discrete device testing)
        wafer = Wafer({
            'number': 0,
            'START_TIME': None,
            'END_TIME': None
        })
        model.add('wafers', wafer)

        # Parse Excel file
        self._parse_excel_file(infile, model, wafer, header)

        # Log parsed metadata
        self.logger.INFO(
            f"LOT={header.LOT}--DEVICE={header.PRODUCT}--PROGRAM={header.RECIPE}--"
            f"RECIPE_REVISION={header.RECIPE_REVISION}--TIME={header.START_TIME}"
        )

        self.logger.INFO(
            f"Parsing complete: {len(model.tests)} tests, {len(wafer.dies)} dies, "
            f"{len(wafer.sbins)} sbins, {len(wafer.hbins)} hbins"
        )

        return model

    def _parse_excel_file(self, infile: str, model: Model, wafer: Wafer,
                          header: Metadata) -> None:
        """
        Parse Excel file using three-phase approach:
        1. Find "No" marker and parse metadata headers
        2. Parse test headers (rows before "No" marker)
        3. Parse die data (rows after "No" marker)

        Args:
            infile: Path to Excel file
            model: Model object to populate
            wafer: Wafer object to populate
            header: Metadata object to populate
        """
        try:
            workbook = load_workbook(infile, data_only=True, read_only=True)
            worksheet = workbook.worksheets[0]
        except Exception as e:
            error_msg = f"Failed to load Excel file: {e}"
            self.logger.ERROR(error_msg)
            Util.dp_exit(1, pplogger=self.pplogger, error=error_msg)

        # Convert all rows to list for indexed access
        all_rows = list(worksheet.iter_rows(values_only=True))
        
        # Parse state
        raw_header: Dict[str, str] = {}
        test_names: List[str] = []
        test_numbers: List[str] = []
        lo_limits: List[str] = []
        hi_limits: List[str] = []
        units: List[str] = []
        
        # PHASE 1: Find the "No" marker and parse metadata
        test_start_row = None
        for row_idx, row in enumerate(all_rows):
            if not row:
                continue
                
            col_a = self._clean_cell(row[0]) if row else ''
            
            # Check for "No" marker in column A
            if col_a and self._PATTERNS['test_table_marker'].match(col_a):
                self.logger.INFO(f"Found 'No' marker at row {row_idx + 1}")
                test_start_row = row_idx
                break
                
            # Parse known header fields (column A has label)
            if col_a in self._HEADER_LABELS:
                col_b = self._clean_cell(row[1]) if len(row) > 1 else ''
                col_c = self._clean_cell(row[2]) if len(row) > 2 else ''
                value = col_c if col_c else col_b
                value = Util.trim(value) if value else 'NA'
                if self._PATTERNS['whitespace'].match(str(value)):
                    value = 'NA'
                raw_header[col_a] = value
                self.logger.INFO(f"Parsed header: {col_a}={value}")
        
        if test_start_row is None:
            self.logger.WARN("No 'No' marker found - cannot parse test data")
            header._raw = raw_header
            header.LOT = raw_header.get('LotID', 'NA')
            return
        
        # PHASE 2: Parse test headers (look backwards from "No" marker)
        # Test headers should be in the rows immediately before the "No" marker
        search_start = max(0, test_start_row - 10)
        for row_idx in range(search_start, test_start_row):
            row = all_rows[row_idx]
            if not row or len(row) < 2:
                continue
                
            col_a = self._clean_cell(row[0])
            col_b = self._clean_cell(row[1])
            
            # Skip metadata header rows
            if col_a in self._HEADER_LABELS:
                continue
            
            # Look for test headers in column B
            # Test# row - data starts at column C (index 2)
            if col_b and self._PATTERNS['test_num_row'].match(col_b):
                test_numbers = [self._clean_cell(c) for c in row[2:]]
                self.logger.INFO(f"Found Test# row at {row_idx + 1}: {test_numbers}")
                continue
            
            # Test Parameter row - data starts at column C (index 2)
            if col_b and self._PATTERNS['test_param'].match(col_b):
                test_names = [self._clean_cell(c) for c in row[2:]]
                self.logger.INFO(f"Found Test Parameter row at {row_idx + 1}: {test_names}")
                continue
            
            # LL row - data starts at column C (index 2)
            if col_b and self._PATTERNS['ll_limit'].match(col_b):
                lo_limits = [self._clean_cell(c) for c in row[2:]]
                self.logger.INFO(f"Found LL row at {row_idx + 1}: {lo_limits}")
                continue
            
            # HL row - data starts at column C (index 2)
            if col_b and self._PATTERNS['hl_limit'].match(col_b):
                hi_limits = [self._clean_cell(c) for c in row[2:]]
                self.logger.INFO(f"Found HL row at {row_idx + 1}: {hi_limits}")
                continue
            
            # Unit row - data starts at column C (index 2)
            if col_b and self._PATTERNS['unit_row'].match(col_b):
                units = [self._clean_cell(c) for c in row[2:]]
                self.logger.INFO(f"Found Unit row at {row_idx + 1}: {units}")
                continue
        
        # PHASE 3: Parse die data (rows after "No" marker)
        for row_idx in range(test_start_row + 1, len(all_rows)):
            row = all_rows[row_idx]
            if not row:
                continue
            
            row_data = [self._clean_cell(cell) for cell in row]
            col_a = row_data[0] if row_data else ''
            
            # Data row: col_a is numeric (die index)
            if col_a and self._PATTERNS['numeric_row'].match(col_a):
                self._parse_die_data(row_data, wafer, test_names)
        
        # Set metadata
        header._raw = raw_header
        header.LOT = raw_header.get('LotID', 'NA')
        
        # Create Test objects
        if test_names:
            test_count = 0
            for i, name in enumerate(test_names):
                if name and not self._PATTERNS['whitespace'].match(name):
                    test_count += 1
                    lsl = lo_limits[i] if i < len(lo_limits) else ''
                    hsl = hi_limits[i] if i < len(hi_limits) else ''
                    unit = units[i] if i < len(units) else ''
                    test_num_raw = test_numbers[i] if i < len(test_numbers) else ''
                    test_num = self._clean_cell(test_num_raw) if test_num_raw else str(test_count)

                    test = Test({
                        'number': test_num,
                        'name': name,
                        'units': unit.strip() if unit else '',
                        'LPL': lsl,
                        'HPL': hsl,
                        'LSL': lsl,
                        'HSL': hsl,
                        'LOL': '',
                        'HOL': '',
                        'LWL': '',
                        'HWL': ''
                    })

                    model.add('tests', test)
                    wafer.add('tests', test)
        else:
            self.logger.WARN("No test names found - skipping test creation")

    def _clean_cell(self, value: Any) -> str:
        """Clean cell value: handle None, convert to string, strip whitespace."""
        if value is None or str(value).lower() == 'nan':
            return ''
        return str(value).strip()

    def _parse_die_data(self, row_data: List[str], wafer: Wafer,
                        test_names: List[str]) -> None:
        """
        Parse one data row into a Die object.
        
        Column structure: No | BIN | blank | test1 | test2 | ...
        """
        if len(row_data) < 2:
            return

        die_index = row_data[0]
        soft_bin = row_data[1]
        bin_numeric = re.sub(r'\D', '', soft_bin) or '0'
        pass_fail = 'P' if bin_numeric == '1' else 'F'

        die = Die({
            'partid': die_index,
            'soft_bin': bin_numeric,
            'hard_bin': bin_numeric,
            'bindesc': f"SWBin_{bin_numeric.zfill(3)}",
            'site': 1,
            'touchdown_num': -1,
            'ecid': die_index,
        })

        # Test results start at column C (index 2)
        for i in range(len(test_names)):
            col_idx = 2 + i
            if col_idx < len(row_data):
                result_val = row_data[col_idx]
                cleaned = self._clean_cell(result_val)
                cleaned = re.sub(r'(Over|undef)', '', cleaned, flags=re.IGNORECASE).strip()
                die.add('result', Util.rep_na(cleaned))
            else:
                die.add('result', 'NA')

        wafer.add('dies', die)

        # Track bins
        for bin_type, bin_list in [('sbins', 'SWBin'), ('hbins', 'HWBin')]:
            bin_name = f"{bin_list}_{bin_numeric.zfill(3)}"
            bin_obj = wafer.find(bin_type, {"number": bin_numeric})
            if bin_obj is None:
                wafer.add(bin_type, Bin({
                    'number': bin_numeric,
                    'name': bin_name,
                    'bindesc': bin_name,
                    'PF': pass_fail,
                    'count': 1,
                }))
            else:
                bin_obj.count += 1
